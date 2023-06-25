import numpy as np
import openpyxl
import os
from math import radians, cos, sin, asin, sqrt
from datetime import datetime, timedelta

def haversine_distance(lon1, lat1, lon2, lat2):
    # Convert decimal degrees to radians
    lon1, lat1, lon2, lat2 = map(radians, [lon1, lat1, lon2, lat2])

    # Haversine formula
    dlon = lon2 - lon1
    dlat = lat2 - lat1
    a = sin(dlat/2)**2 + cos(lat1) * cos(lat2) * sin(dlon/2)**2
    c = 2 * asin(sqrt(a))
    r = 6371  # Radius of the Earth in kilometers
    return c * r


def idw_interpolation(known_locations, known_values, unknown_location, power_parameter):
    # distances = np.array([haversine_distance(*known_loc, *unknown_location) for known_loc in known_locations])
  distances = []
  known = []
  c = -1
  for known_loc in known_locations:
    c = c + 1
    if known_values[c] is not None:
      # print(known_locations)
      # print(known_values[c])
      distance = haversine_distance(*known_loc, *unknown_location)
      distances.append(distance)
      known.append(known_values[c])

  distances = np.array(distances)
  known = np.array(known)
  # print(known)
  weights = 1 / (distances ** power_parameter)
  estimated_value = np.sum(weights * known) / np.sum(weights)
  return estimated_value

def fillRow(known, unknown, uPos, is_row, row, var):
    l = len(known)
    latLon = []
    dis = []
    power_parameter = 2

    variables = {
        "p": [],
        "wd": [],
        "ws": [],
        "h": [],
        "r": [],
        "t": []
    }

    lat_index = 3
    lon_index = 4
    variables_index = {
        "p": 5,
        "wd": 6,
        "ws": 7,
        "h": 8,
        "r": 9,
        "t": 10
    }

    for k in range(l):
        lon = float(str(known[k][lon_index]).replace(',', '.'))
        lat = float(str(known[k][lat_index]).replace(',', '.'))
        latLon.append([lon, lat])
        # useLocation = [True,True,True,True,True,True]
        i = 0
        if is_row:
          for variable, index in variables_index.items():
            i = i + 1
            if known[k][index] is not None:
                variables[variable].append(float(str(known[k][index]).replace(',', '.')))
            else: variables[variable].append(None)
        else:
          for variable, index in variables_index.items():
            if var == variable and known[k][index] is not None:
                variables[variable].append(float(str(known[k][index]).replace(',', '.')))
            else: variables[variable].append(None)

    if is_row:
        interpolated_values = []
        for variable in variables_index.keys():
            if variables[variable]:
                interpolated_value = round(idw_interpolation(latLon, variables[variable], uPos, power_parameter), 1)
                interpolated_values.append(interpolated_value)
            else:
                interpolated_values.append(None)


        new_row = (row[0], row[1], known[0][2], row[3], row[4]) + tuple(interpolated_values)
        return new_row
    else:
        if variables[var]:
            return round(idw_interpolation(latLon, variables[var], uPos, power_parameter), 1)


def getKnown(data, city, time, stations):
  known = []
  # print(stations)
  for station in stations:
    # print(station)
    if station != city:
      city_data = data.get(station)
      # print(city_data)
      if city_data :
        filtered_data = [d for d in city_data if d != None and d[2] == time]
        # print(filtered_data)
        if filtered_data:
            known.append(filtered_data[0])
  return known


def process(city, data, stations):
    # workbook = openpyxl.load_workbook(checkFile)
    date_format = "%Y-%m-%d %H:%M:%S"
    # sheet_names = workbook.sheetnames
    varName=["","","","","","p","wd","ws","h","r","t"]
    headers = ['ClimNo','StasName','DateT','Latitude','Longitude','Pressure','WindDir','WindSpeed','Humidity','Rain','Temperature']

    # Iterate over the sheets
    # for sheet_name in sheet_names:
        # sheet = workbook[sheet_name]
    print(f"Processing station: {city}")

    prev_date = datetime.strptime("2010-01-01 00:00:00", date_format)
        # end_date = datetime.strptime("2010-01-01 18:00:00", date_format)
    one_hour = timedelta(hours=1)
        # missCount = 0
        # row_num = -1
        # missing_rows = []
    prev_row = ""
    # print(data[city])
    for index,rows in enumerate(data[city]):
      date = prev_date + one_hour
      # date = datetime.strptime(dt, date_format)
      # print(rows)
      if rows == None:
        # time_difference = date - prev_date

        # if time_difference > one_hour:
          # hours = int(time_difference.total_seconds() / 3600)
        print("Inserting missing data at: " + prev_date.strftime(date_format))
          # for _ in range(1, hours):
            # prev_date += one_hour
        known = getKnown(data, city, prev_date, stations.keys())
            # print(known)
        newRow = fillRow(known, city, stations[city], True, prev_row, False)
        data[city][index] = newRow
      else:
        if any(v is None for v in rows[5:11]):
          print("Inserting missing variable(s) at : " +  prev_date.strftime(date_format))
          known = getKnown(data, city, prev_date, stations.keys())
          new_row = list(rows)
          for i in range(5, 11):
            if new_row[i] is None:
              new_row[i] = fillRow(known, city, stations[city], False, prev_row, varName[i])
          data[city][index] = new_row
      prev_date = date
      prev_row = data[city][index]
    # print(data)

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(headers)
    print("Writing to station: " + city)
    for row_index, row in enumerate(data[city], start=2):
      for column_index, value in enumerate(row, start=1):
        sheet.cell(row=row_index, column=column_index, value=value)

    output_file = "test_results/WC_" + city + ".xlsx"
    workbook.save(output_file)

def storeData(file_path, stData, stations):
    workbook = openpyxl.load_workbook(file_path, read_only=True)
    sheet_names = workbook.sheetnames
    date_format = "%Y-%m-%d %H:%M:%S"
    gotLat = False
    # Iterate over the sheets

    prev_date = datetime.strptime("2010-01-01 00:00:00", date_format)
    end_date = datetime.strptime("2010-01-01 18:00:00", date_format)
    one_hour = timedelta(hours=1)

    for sheet_name in sheet_names:
        sheet = workbook[sheet_name]
        print(f"Getting data from Sheet Name: {sheet_name} in file: {file_path}")

        for row in sheet.iter_rows(values_only=True):
          city = row[1]
          # print(row)
          if gotLat == False and city != "StasName":
            lon = float(str(row[4]).replace(',', '.'))
            lat = float(str(row[3]).replace(',', '.'))
            # lonLat = [lon, lat]
            stations[city]= []
            stations[city].append(lon)
            stations[city].append(lat)
            gotLat = True

          try:
            dt = str(row[2])
            date = datetime.strptime(dt, date_format)
            time_difference = date - prev_date

            if city != "StasName" and city != None:

              if time_difference > one_hour:
                hours = int(time_difference.total_seconds() / 3600)
                for c in range(1, hours):
                  # if city != "StasName" and city != None:
                  # print(city)
                  if city not in stData:
                    stData[city] = []
                  stData[city].append(None)
              # else:
              # if city != "StasName" and city != None:
              if city not in stData:
                stData[city] = []
              stData[city].append(row)
          except ValueError as e:
            continue
          prev_date = date

    # print(stData)
    return stData, stations

  
def main():
    print("Starting")
    cutoff = 0.1
    headers = ['ClimNo','StasName','DateT','Latitude','Longitude','Pressure','WindDir','WindSpeed','Humidity','Rain','Temperature']
    # checkFile = "PAARL.xlsx"
    # print(getLatLon(checkFile))#

    # Specify the folder path
    folder_path = "test"  # Replace with your folder path

    Stations = {}
    stationPos = []
    output = "missing.txt"
    data = {}
    # Iterate through files in the folder
    # index = 0
    for filename in os.listdir(folder_path):
      file_path = os.path.join(folder_path, filename)
      if os.path.isfile(file_path):
       
        data,Stations = storeData(file_path,data,Stations)
       
    print("Data stored and starting filling missing data")
   

    print("-------------------------------------------")
    
    for city in data:
      # print(data[city])
      process(city,data,Stations)
      # for r in data[city]:
      #   print(r)
    # print(data)
    print("Completed preprocessing")

if __name__ == '__main__':
    main()
