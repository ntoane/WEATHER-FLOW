import openpyxl

def convert(headers):

    stations = []
    i = 0
    stData = []
    stData.append([])
    # Load the Excel file
    workbook = openpyxl.load_workbook('newData/wc hourly data 2010-2022.xlsx', read_only=True)


    # Get the names of all sheets
    sheet_names = workbook.sheetnames

    # Iterate over the sheets
    for sheet_name in sheet_names:
        sheet = workbook[sheet_name]
        print(f"Sheet Name: {sheet_name}")
        
        # Iterate over all rows in the sheet
        for row in sheet.iter_rows(values_only=True):
            # Print the values of each cell in the row
            if row[1] not in  stations:
                stations.append(row[1])
                stData.append([])
                stData[i].append(row)
                i = i+1
            else :
                index = stations.index(row[1])
                stData[index].append(row)

    for x in range(0,len(stations)):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(headers)
        print(stations[x])
        for row_index, row in enumerate(stData[x], start=2):
            for column_index, value in enumerate(row, start=1):
                sheet.cell(row=row_index, column=column_index, value=value)
        
        output_file = f"Western Cape/{str(stations[x])}.xlsx"
        workbook.save(output_file)

    # print(stData)

def main():
    print("Starting")
    headers = ['ClimNo','StasName','DateT','Latitude','Longitude','Pressure','WindDir','WindSpeed','Humidity','Rain','Temperature']

    convert (headers)

    print("Completed filtering")

if __name__ == '__main__':
    main()
