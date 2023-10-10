import openpyxl
from datetime import datetime, timedelta
import os

def process(checkFile,cutoff, cutStations, out):
  # workbook = openpyxl.load_workbook('newData/nc hourly data 2010-2022.xlsx')
    workbook = openpyxl.load_workbook(checkFile, read_only=True)

    date_format = "%Y-%m-%d %H:%M:%S"
    # Get the names of all sheets
    sheet_names = workbook.sheetnames

    # Iterate over the sheets
    for sheet_name in sheet_names:
        sheet = workbook[sheet_name]
        print(f"File Name: {checkFile}")

        # prev = datetime(2010, 1, 1, 0, 0, 0)
        prev_date = "2010-01-01 00:00:00"
        end_date = "2022-12-31 00:00:00"
        # Define a timedelta of one hour
        one_hour = timedelta(hours=1)
        missCount = 0
        # Convert the strings to datetime objects
        prev_date = datetime.strptime(prev_date, date_format)
        end_date = datetime.strptime(end_date, date_format)

        totalCount = ((end_date -prev_date).total_seconds())/3600 + 1
        # Iterate over all rows in the sheet
        for row in sheet.iter_rows(values_only=True):
            
            try:
              # totalCount +=1
              dt =str(row[2])
              date = datetime.strptime(dt, date_format)
              time_difference = date - prev_date
              # print(time_difference)
              # print("----------------------------------")
              if(time_difference>one_hour):
                # print(prev_date)
                # print(date)
                hours = (time_difference.total_seconds())/3600
                #print(hours)
                missCount = missCount + hours
                # print(row[5])
              prev_date = date
            except ValueError as e:
              #print(f"Encountered a Error: {e}")
              continue

            if(row[5]== None or row[6] == None or row[7]== None or row[8] == None or row[9]== None or row[10] == None):
              missCount = missCount + 1
	
        #print(missCount)
        missing = missCount/totalCount
        if (missing>cutoff):
          # tooMuch = True
          print("------------------------------------------------")
          print("Exceeding cutoff" + checkFile)
          print("------------------------------------------------")
          cutStations.append(checkFile)
        # print(totalCount)

        with open(out, 'a') as file:
          file.write(checkFile + ":   " + str(missing) + "\n")

    return cutStations
           

def main():
    print("Starting")
    cutoff = 0.1
   
    # Specify the folder path
    folder_path = "dataset"  # Replace with your folder path

    cutStations = []
    output = "dataset.txt"

    # Iterate through files in the folder
    for filename in os.listdir(folder_path):
      file_path = os.path.join(folder_path, filename)
      if os.path.isfile(file_path):
        cutStations = process(file_path,cutoff,cutStations,output)

    print(cutStations)
    print("Completed preprocessing")

if __name__ == '__main__':
    main()
